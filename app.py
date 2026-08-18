from flask import Flask, render_template, request, jsonify
from datetime import datetime
import os
from openpyxl import load_workbook

app = Flask(__name__)

# ── Configuration ─────────────────────────────────────────────────────────────
EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "PAYROLL.xlsx")

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# ── Load payroll from Excel ───────────────────────────────────────────────────
def _safe_float(value) -> float:
    """Convert an Excel cell value to float, ignoring non-numeric values."""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    return 0.0

def load_payroll():
    """Read every payroll row from data/PAYROLL.xlsx.

    Expected columns:
      A: PERIOD START, B: PERIOD END, C: DATE PAID, D: HOURS,
      E: EMPLOYEE NAME, F: POSITION, G: PAY RATE/HR, H: AMOUNT, I: COMMENTS
    """
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    records = []
    if ws is None:
        return records
    for row in ws.iter_rows(min_row=2, values_only=True):
        period_start, period_end, date_paid, hours, name, position, rate, amount, comments = row[:9]
        if not name or str(name).strip().upper() == "GRAND TOTAL":
            continue
        # Skip rows where the date cells are formulas without cached values
        if not (isinstance(period_start, datetime)
                and isinstance(period_end, datetime)
                and isinstance(date_paid, datetime)):
            continue

        hours    = _safe_float(hours)
        rate     = _safe_float(rate)
        position = str(position).strip() if position else "Staff"
        # AMOUNT may be a formula (=D*F) → use cached value, else compute it
        if isinstance(amount, (int, float)) and not isinstance(amount, bool):
            amount_val = round(float(amount), 2)
        else:
            amount_val = round(hours * rate, 2)

        records.append({
            "emp_name":     str(name).strip(),
            "position":     position,
            "pay_rate":     rate,
            "hours":        hours,
            "amount":       amount_val,
            "month":        period_end.strftime("%b"),
            "month_num":    period_end.month,
            "period_start": period_start.strftime("%Y-%m-%d"),
            "period_end":   period_end.strftime("%Y-%m-%d"),
            "date_paid":    date_paid.strftime("%Y-%m-%d"),
            "comments":     str(comments).strip() if comments else "",
        })
    return records

# ── Build employees from the payroll rows ─────────────────────────────────────
def build_employees(payroll):
    emp_map = {}
    for r in payroll:
        name = r["emp_name"]
        if name not in emp_map:
            emp_map[name] = {
                "id":         len(emp_map) + 1,
                "name":       name,
                "position":   r["position"],
                "pay_rate":   r["pay_rate"],
                "status":     "Active",
                "records":    [],
            }
        emp_map[name]["records"].append(r)
        emp_map[name]["pay_rate"] = r["pay_rate"]  # keep the latest rate
    for e in emp_map.values():
        e["records"].sort(key=lambda x: (x["month_num"], x["period_start"]))
    return list(emp_map.values())

def get_available_months(payroll):
    return sorted({r["month"] for r in payroll}, key=MONTHS.index)

def get_available_positions(payroll):
    return sorted({r["position"] for r in payroll})

def get_filtered_data(payroll, month=None, position=None, search=None):
    emps = build_employees(payroll)
    if position and position != "All":
        emps = [e for e in emps if e["position"] == position]
    if search:
        q = search.lower()
        emps = [e for e in emps
                if q in e["name"].lower()
                or q in e["position"].lower()]

    rows = []
    for e in emps:
        for r in e["records"]:
            if month and month != "All" and r["month"] != month:
                continue
            rows.append({**e, **r, "records": None})
    return emps, rows

# ── Test-data cleanup (moved from cleanup_test.py) ─────────────────────────────
def cleanup_test_data():
    """Remove test/scratch data from the payroll workbook.

    Idempotent: safe to run on every startup. Clears the scratch row (1001)
    and reports any leftover \"Test Sync\" employee rows for awareness.
    """
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    if ws is None:
        return

    # Force-clear the scratch row (1001) across all columns
    cleared = any(
        ws.cell(row=1001, column=c).value is not None
        for c in range(1, 30)
    )
    for c in range(1, 30):
        ws.cell(row=1001, column=c).value = None

    # Report leftover test employees
    test_emps = sorted({
        str(row[4]) for row in ws.iter_rows(min_row=2, values_only=True)
        if row[4] and "Test Sync" in str(row[4])
    })

    try:
        wb.save(EXCEL_PATH)
    except OSError:
        # Filesystem may be read-only (e.g. Vercel serverless) — nothing to
        # persist in that case since cleanup is only relevant on a writable FS.
        return

    if cleared:
        print("cleanup_test_data: cleared scratch row 1001")
    if test_emps:
        print(f"cleanup_test_data: found test employees -> {test_emps}")

# ── Routes ────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    payroll = load_payroll()
    return render_template("index.html",
        months=get_available_months(payroll),
        positions=get_available_positions(payroll))

@app.route("/api/summary")
def api_summary():
    month    = request.args.get("month",  "All")
    position = request.args.get("position", "All")
    search   = request.args.get("search", "")

    # Reload payroll from Excel on every request so the dashboard stays in sync
    payroll = load_payroll()
    available_months = get_available_months(payroll)
    emps, rows = get_filtered_data(payroll, month, position, search)

    total_payroll = round(sum(r["amount"] for r in rows), 2)
    total_hours   = round(sum(r["hours"]  for r in rows), 2)
    active_emps   = len([e for e in emps if e["status"] == "Active"])
    avg_rate      = round(sum(e["pay_rate"] for e in emps) / len(emps), 2) if emps else 0
    on_leave      = len([e for e in emps if e["status"] == "On Leave"])

    cards = {
        "total_payroll":    total_payroll,
        "total_employees":  len(emps),
        "active_employees": active_emps,
        "avg_pay_rate":     avg_rate,
        "total_hours":      total_hours,
        "on_leave":         on_leave,
    }

    # Pie – payroll by position
    by_pos = {}
    for r in rows:
        by_pos[r["position"]] = by_pos.get(r["position"], 0) + r["amount"]
    pie = {"labels": list(by_pos.keys()), "values": [round(v, 2) for v in by_pos.values()]}

    # Bar – top 10 employees by total pay
    by_emp = {}
    for r in rows:
        by_emp[r["name"]] = by_emp.get(r["name"], 0) + r["amount"]
    top_emps = sorted(by_emp.items(), key=lambda x: x[1], reverse=True)[:10]
    bar = {"labels": [x[0] for x in top_emps], "values": [round(x[1], 2) for x in top_emps]}

    # Line – monthly payroll trend (only months present in the data)
    monthly = {}
    for r in rows:
        monthly[r["month"]] = monthly.get(r["month"], 0) + r["amount"]
    ordered = {m: round(monthly.get(m, 0), 2) for m in available_months}
    line = {"labels": list(ordered.keys()), "values": list(ordered.values())}

    # Top 5 highest paid employees
    top5 = [{"name": x[0], "amount": round(x[1], 2)} for x in
            sorted(by_emp.items(), key=lambda x: x[1], reverse=True)[:5]]

    # Table rows – match the PAYROLL.xlsx columns
    table = []
    for e in emps:
        emp_rows = [r for r in e["records"] if month == "All" or r["month"] == month]
        for r in emp_rows:
            table.append({
                "period_start": r["period_start"],
                "period_end":   r["period_end"],
                "date_paid":    r["date_paid"],
                "hours":        r["hours"],
                "name":         e["name"],
                "position":     e["position"],
                "pay_rate":     r["pay_rate"],
                "amount":       r["amount"],
                "comments":     r["comments"],
            })

    return jsonify(cards=cards, pie=pie, bar=bar, line=line, top5=top5, table=table)

if __name__ == "__main__":
    cleanup_test_data()
    app.run(debug=True, port=5000)